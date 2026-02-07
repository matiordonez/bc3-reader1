# -*- coding: utf-8 -*-
"""Exportador de presupuestos BC3 a Excel (XLSX)."""

import io
from pathlib import Path
from typing import List, Dict, Union

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter


def export_to_xlsx(partidas: List[Dict], output_path: str, titulo: str = "Presupuesto BC3") -> str:
    """
    Exporta las partidas del presupuesto a un archivo Excel.
    
    Args:
        partidas: Lista de diccionarios con los datos de cada partida
        output_path: Ruta del archivo de salida .xlsx
        titulo: Título del documento
        
    Returns:
        Ruta del archivo generado
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Presupuesto"
    
    # Estilos
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Título
    ws.merge_cells('A1:H1')
    ws['A1'] = titulo
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.row_dimensions[1].height = 25
    
    # Encabezados
    headers = ['Código', 'Descripción', 'Unidad', 'Cantidad', 'Precio Unit.', 'Importe', 'Texto', 'Descomposición']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = thin_border
    
    # Datos
    for row_idx, partida in enumerate(partidas, start=4):
        ws.cell(row=row_idx, column=1, value=partida.get('codigo', ''))
        ws.cell(row=row_idx, column=2, value=partida.get('descripcion', ''))
        ws.cell(row=row_idx, column=3, value=partida.get('unidad', ''))
        ws.cell(row=row_idx, column=4, value=partida.get('cantidad', ''))
        ws.cell(row=row_idx, column=5, value=partida.get('precio_unitario', ''))
        ws.cell(row=row_idx, column=6, value=partida.get('importe', ''))
        ws.cell(row=row_idx, column=7, value=partida.get('texto_largo', ''))
        ws.cell(row=row_idx, column=8, value=partida.get('descomposicion', ''))
        
        for col in range(1, 9):
            ws.cell(row=row_idx, column=col).border = thin_border
            ws.cell(row=row_idx, column=col).alignment = Alignment(wrap_text=True, vertical='top')
    
    # Ajustar anchos de columna
    column_widths = [15, 40, 10, 12, 15, 15, 50, 40]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    # Guardar
    output_path = Path(output_path)
    if output_path.suffix.lower() != '.xlsx':
        output_path = output_path.with_suffix('.xlsx')
    
    wb.save(str(output_path))
    return str(output_path)


def export_to_xlsx_bytes(partidas: List[Dict], titulo: str = "Presupuesto BC3") -> bytes:
    """Exporta a Excel y retorna los bytes (para respuestas HTTP)."""
    buffer = io.BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "Presupuesto"
    
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    ws.merge_cells('A1:H1')
    ws['A1'] = titulo
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.row_dimensions[1].height = 25
    
    headers = ['Código', 'Descripción', 'Unidad', 'Cantidad', 'Precio Unit.', 'Importe', 'Texto', 'Descomposición']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = thin_border
    
    for row_idx, partida in enumerate(partidas, start=4):
        ws.cell(row=row_idx, column=1, value=partida.get('codigo', ''))
        ws.cell(row=row_idx, column=2, value=partida.get('descripcion', ''))
        ws.cell(row=row_idx, column=3, value=partida.get('unidad', ''))
        ws.cell(row=row_idx, column=4, value=partida.get('cantidad', ''))
        ws.cell(row=row_idx, column=5, value=partida.get('precio_unitario', ''))
        ws.cell(row=row_idx, column=6, value=partida.get('importe', ''))
        ws.cell(row=row_idx, column=7, value=partida.get('texto_largo', ''))
        ws.cell(row=row_idx, column=8, value=partida.get('descomposicion', ''))
        for col in range(1, 9):
            ws.cell(row=row_idx, column=col).border = thin_border
            ws.cell(row=row_idx, column=col).alignment = Alignment(wrap_text=True, vertical='top')
    
    column_widths = [15, 40, 10, 12, 15, 15, 50, 40]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    wb.save(buffer)
    return buffer.getvalue()
